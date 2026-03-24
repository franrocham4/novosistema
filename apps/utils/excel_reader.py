import os
import warnings
from datetime import datetime, date

import openpyxl
from django.conf import settings


def _get_workbook():
    path = settings.EXCEL_FILE_PATH
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Arquivo Excel não encontrado: {path}\n"
            "Configure a variável de ambiente EXCEL_FILE_PATH com o caminho correto."
        )
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path, keep_vba=True, data_only=True)
    return wb


def _fmt_date(value):
    if isinstance(value, (datetime, date)):
        return value.strftime('%d/%m/%Y') if hasattr(value, 'strftime') else str(value)
    return value or ''


def _fmt_number(value):
    if isinstance(value, (int, float)):
        return value
    return 0


def get_liquidacao_data():
    wb = _get_workbook()
    ws = wb['LIQUIDAÇÃO 2025']

    header_row = 5  # row 5 has headers (1-indexed)
    data_start = 6

    records = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not any(v is not None for v in row[:10]):
            continue
        records.append({
            'data_liquid': _fmt_date(row[0]),
            'cod_empresa': row[1] or '',
            'empresa': row[2] or '',
            'processo': row[3] or '',
            'contrato': row[4] or '',
            'empenho': row[5] or '',
            'nf': row[6] or '',
            'mes_desembolso': row[7] or '',
            'periodo_inicial': _fmt_date(row[8]),
            'periodo_final': _fmt_date(row[9]),
            'valor_bruto': _fmt_number(row[10]),
            'valor_retencao': _fmt_number(row[11]),
            'valor_liquido': _fmt_number(row[12]),
            'vencimento': _fmt_date(row[13]),
            'valores_pagos': _fmt_number(row[14]),
            'pagto': _fmt_date(row[15]),
            'status': row[16] or '',
            'dias_atraso': row[17] or '',
            'previsao_pagto': _fmt_date(row[18]),
        })

    total_bruto = sum(r['valor_bruto'] for r in records)
    total_retencao = sum(r['valor_retencao'] for r in records)
    total_liquido = sum(r['valor_liquido'] for r in records)
    total_pago = sum(r['valores_pagos'] for r in records if r['status'] == 'PAGO')
    em_atraso = sum(r['valor_liquido'] for r in records if r['status'] == 'ATRASADO')
    a_pagar = sum(r['valor_liquido'] for r in records if r['status'] not in ('PAGO',))

    status_counts = {}
    for r in records:
        s = r['status'] or 'N/A'
        status_counts[s] = status_counts.get(s, 0) + 1

    return {
        'records': records,
        'totais': {
            'total_bruto': total_bruto,
            'total_retencao': total_retencao,
            'total_liquido': total_liquido,
            'total_pago': total_pago,
            'em_atraso': em_atraso,
            'a_pagar': a_pagar,
        },
        'status_counts': status_counts,
        'ultima_atualizacao': _get_file_mtime(),
    }


def get_contratos_data():
    wb = _get_workbook()
    ws = wb['ALT. ORÇ. E PROJ.']

    header_row = 5
    data_start = 6

    records = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not any(v is not None for v in row[:9]):
            continue
        records.append({
            'programa_anulado': row[1] or '',
            'projeto_anulado': row[2] or '',
            'ficha_anulado': row[3] or '',
            'programa_suplementado': row[4] or '',
            'projeto_suplementado': row[5] or '',
            'ficha_suplementado': row[6] or '',
            'valor': _fmt_number(row[7]),
            'justificativa': row[8] or '',
        })

    total_valor = sum(r['valor'] for r in records)

    return {
        'records': records,
        'total_valor': total_valor,
        'ultima_atualizacao': _get_file_mtime(),
    }


def get_empenho_data():
    wb = _get_workbook()
    ws = wb['EMPENHO ']

    header_row = 3
    data_start = 4

    records = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not any(v is not None for v in row[:9]):
            continue
        records.append({
            'data': _fmt_date(row[0]),
            'programa': row[1] or '',
            'ficha': row[2] or '',
            'acao': row[3] or '',
            'projeto': row[4] or '',
            'reserva': row[5] or '',
            'cod_empresa': row[6] or '',
            'empresa': row[7] or '',
            'empenho': row[8] or '',
            'valor_empenho': _fmt_number(row[9]),
            'estorno': _fmt_number(row[10]),
            'saldo_empenho': _fmt_number(row[11]),
            'liquidado': _fmt_number(row[12]),
            'saldo': _fmt_number(row[13]),
        })

    total_empenho = sum(r['valor_empenho'] for r in records)
    total_liquidado = sum(r['liquidado'] for r in records)
    total_saldo = sum(r['saldo'] for r in records)
    total_estorno = sum(r['estorno'] for r in records)

    return {
        'records': records,
        'totais': {
            'total_empenho': total_empenho,
            'total_liquidado': total_liquidado,
            'total_saldo': total_saldo,
            'total_estorno': total_estorno,
        },
        'ultima_atualizacao': _get_file_mtime(),
    }


def get_dashboard_data():
    liq = get_liquidacao_data()
    emp = get_empenho_data()
    return {
        'liquidacao': liq,
        'empenho': emp,
        'ultima_atualizacao': _get_file_mtime(),
    }


def _get_file_mtime():
    path = settings.EXCEL_FILE_PATH
    if os.path.exists(path):
        ts = os.path.getmtime(path)
        return datetime.fromtimestamp(ts).strftime('%d/%m/%Y %H:%M:%S')
    return 'N/A'
